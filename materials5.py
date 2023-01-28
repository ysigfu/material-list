import os
import openpyxl

# PROGRAM TO GET MATERIAL LISTS IN EACH TAKEOFF FILE IN ONE SPREADSHEET
# Declaring locations
directory = '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4'
#folder = '/home/abramnaranjo/.bkp/knells/land'
#writetofile = '/home/abramnaranjo/.bkp/knells/material-list.csv'
# For each file in folder $path
RCtotal = []
FIRECODEtotal = []
MOLDTOUGHtotal = []
FIRECODECtotal = []
MUDtotal = []
TAPEtotal = []
ANGLEtotal = []
MAINStotal = []
TEEStotal = []
PANHEADStotal = []
PINtotal = []
SCREWStotal = []
metal = []
track = []
highimpact = []
FURRINGChannel = []


for filename in os.listdir(directory):
    path = os.path.join(directory, filename)
    wb_obj = openpyxl.load_workbook(path)
    # Makes "Setup" the active sheet in takeoff program
    sh = wb_obj['Setup']
    # Gets the project name from cell B1 in the Setup excel spreadsheet
    project_name = str(sh.cell(row=1, column=2).value)
    # Changes the active spreadsheet to "Material"
    sh = wb_obj['Material']
    # Cell Range A1:F12 on "Materials"
   # if path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/A16K.xlsm':
   #     multiply = 2
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/A20K.xlsm':
   #     multiply = 1
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/A28A.xlsm':
   #     multiply = 1
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/A28.xlsm':
   #     multiply = 2
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/A28S.xlsm':
   #     multiply == 8
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B28K.xlsm':
   #     multiply = 2
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B30E.xlsm':
   #     multiply = 5
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B30K.xlsm':
   #     multiply = 1
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B34.xlsm':
   #     multiply = 1
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B36.xlsm':
   #     multiply = 3
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B36S.xlsm':
   #     multiply = 2
   # #elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/UNITS/B38A.xlsm':
   # #    multiply = 2
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B38.xlsm':
   #     multiply = 6
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B38S.xlsm':
   #     multiply = 10
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/B46D.xlsm':
   #     multiply = 4
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/C40E.xlsm':
   #     multiply = 3
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/C40KA.xlsm':
   #     multiply = 1
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/C40K.xlsm':
   #     multiply = 4
   # elif path == '/home/abramnaranjo/Desktop/WINDSONG/windsong-apartments/TAKEOFFS/LEVEL4/C46.xlsm':
   #     multiply = 2

 #   print(project_name)
    cellA1 = sh.cell(row=1, column=1).value
    cellA2 = sh.cell(row=2, column=1).value
    cellA3 = sh.cell(row=3, column=1).value
    cellA4 = sh.cell(row=4, column=1).value
    cellA5 = sh.cell(row=5, column=1).value
    cellA6 = sh.cell(row=6, column=1).value
    cellA7 = sh.cell(row=7, column=1).value
    cellA8 = sh.cell(row=8, column=1).value
    cellA9 = sh.cell(row=9, column=1).value
    cellA10 = sh.cell(row=10, column=1).value
    cellA11 = sh.cell(row=11, column=1).value
    cellA12 = sh.cell(row=12, column=1).value
    cellA13 = sh.cell(row=13, column=1).value
    cellB1 = sh.cell(row=1, column=2).value
    cellB2 = sh.cell(row=2, column=2).value
    cellB3 = sh.cell(row=3, column=2).value
    cellB4 = sh.cell(row=4, column=2).value
    cellB5 = sh.cell(row=5, column=2).value
    cellB6 = sh.cell(row=6, column=2).value
    cellB7 = sh.cell(row=7, column=2).value
    cellB8 = sh.cell(row=8, column=2).value
    cellB9 = sh.cell(row=9, column=2).value
    cellB10 = sh.cell(row=10, column=2).value
    cellB11 = sh.cell(row=11, column=2).value
    cellB12 = sh.cell(row=12, column=2).value
    cellB13 = sh.cell(row=13, column=2).value
    cellC1 = sh.cell(row=1, column=3).value
    cellC2 = sh.cell(row=2, column=3).value
    cellC3 = sh.cell(row=3, column=3).value
    cellC4 = sh.cell(row=4, column=3).value
    cellC5 = sh.cell(row=5, column=3).value
    cellC6 = sh.cell(row=6, column=3).value
    cellC7 = sh.cell(row=7, column=3).value
    cellC8 = sh.cell(row=8, column=3).value
    cellC9 = sh.cell(row=9, column=3).value
    cellC10 = sh.cell(row=10, column=3).value
    cellC11 = sh.cell(row=11, column=3).value
    cellC12 = sh.cell(row=12, column=3).value
    cellC13 = sh.cell(row=13, column=3).value
    cellD1 = sh.cell(row=1, column=4).value
    cellD2 = sh.cell(row=2, column=4).value
    cellD3 = sh.cell(row=3, column=4).value
    cellD4 = sh.cell(row=4, column=4).value
    cellD5 = sh.cell(row=5, column=4).value
    cellD6 = sh.cell(row=6, column=4).value
    cellD7 = sh.cell(row=7, column=4).value
    cellD8 = sh.cell(row=8, column=4).value
    cellD9 = sh.cell(row=9, column=4).value
    cellD10 = sh.cell(row=10, column=4).value
    cellD11 = sh.cell(row=11, column=4).value
    cellD12 = sh.cell(row=12, column=4).value
    cellD13 = sh.cell(row=13, column=4).value
    cellE1 = sh.cell(row=1, column=5).value
    cellE2 = sh.cell(row=2, column=5).value
    cellE3 = sh.cell(row=3, column=5).value
    cellE4 = sh.cell(row=4, column=5).value
    cellE5 = sh.cell(row=5, column=5).value
    cellE6 = sh.cell(row=6, column=5).value
    cellE7 = sh.cell(row=7, column=5).value
    cellE8 = sh.cell(row=8, column=5).value
    cellE9 = sh.cell(row=9, column=5).value
    cellE10 = sh.cell(row=10, column=5).value
    cellE11 = sh.cell(row=11, column=5).value
    cellE12 = sh.cell(row=12, column=5).value
    cellE13 = sh.cell(row=13, column=5).value
    cellF1 = sh.cell(row=1, column=6).value
    cellF2 = sh.cell(row=2, column=6).value
    cellF3 = sh.cell(row=3, column=6).value
    cellF4 = sh.cell(row=4, column=6).value
    cellF5 = sh.cell(row=5, column=6).value
    cellF6 = sh.cell(row=6, column=6).value
    cellF7 = sh.cell(row=7, column=6).value
    cellF8 = sh.cell(row=8, column=6).value
    cellF9 = sh.cell(row=9, column=6).value
    cellF10 = sh.cell(row=10, column=6).value
    cellF11 = sh.cell(row=11, column=6).value
    cellF12 = sh.cell(row=12, column=6).value
    cellF13 = sh.cell(row=13, column=6).value

   # if cellA2 == 'Z Channel':
   #     RCqty = cellE2
   #     RCqty = RCqty * multiply
   #     RCtotal.append(RCqty)
   # if cellA3 == '5/8 Firecode':
   #     FIRECODEqty = cellE3
   #     FIRECODEqty = FIRECODEqty * multiply
   #     FIRECODEtotal.append(FIRECODEqty)
   # if cellA4 == '5/8 Mold Tough/Xp':
   #     MOLDTOUGHqty = cellE4
   #     MOLDTOUGHqty = MOLDTOUGHqty * multiply
   #     MOLDTOUGHtotal.append(MOLDTOUGHqty)
   # if cellA5 == '5/8 Firecode C':
   #     FIRECODECqty = cellE5
   #     FIRECODECqty = FIRECODECqty * multiply
   #     FIRECODECtotal.append(FIRECODECqty)
   # if cellA6 == 'Mud':
   #     MUDqty = cellE6
   #     MUDqty = MUDqty * multiply
   #     MUDtotal.append(MUDqty)
   # if cellA7 =='Tape':
   #     TAPEqty = cellE7
   #     TAPEqty = TAPEqty * multiply
   #     TAPEtotal.append(TAPEqty)
   # if cellC8 == '1 1/2"':
   #     ANGLEqty = cellE8
   #     ANGLEqty = ANGLEqty * multiply
   #     ANGLEtotal.append(ANGLEqty)
   # if cellC9 == "12\' MAINS":
   #     MAINSqty = cellE9
   #     MAINSqty = MAINSqty * multiply
   #     MAINStotal.append(MAINSqty)
   # if cellC10 == "4' TEES":
   #     TEESqty = cellE10
   #     TEESqty = TEESqty * multiply
   #     TEEStotal.append(TEESqty)
   # if cellC11 == '8 1/2" PAN':
   #     PANHEADSqty = cellE11
   #     PANHEADSqty = PANHEADSqty * multiply
   #     PANHEADStotal.append(PANHEADSqty)
   # if cellC12 == '3/4" TF PIN W/FUEL':
   #     PINqty = cellE12
   #     PINqty = PINqty * multiply
   #     PINtotal.append(PINqty)
   # if cellA13 == 'Screws':
   #     SCREWSqty = cellE13
   #     SCREWSqty = SCREWSqty * multiply
   #     SCREWStotal.append(SCREWSqty)



#row1 = cellA1, cellB1, cellC1, cellD1, cellE1
#row2 = cellA2, cellB2, cellC2, cellD2, sum(RCtotal)
#row3 = cellA3, cellB3, cellC3, cellD3, sum(FIRECODEtotal)
#row4 = cellA4, cellB4, cellC4, cellD4, sum(MOLDTOUGHtotal)
#row5 = cellA5, cellB5, cellC5, cellD5, sum(FIRECODECtotal)
#row6 = cellA6, cellB6, cellC6, cellD6, sum(MUDtotal)
#row7 = cellA7, cellB7, cellC7, cellD7, sum(TAPEtotal)
#row8 = cellA8, cellB8, cellC8, cellD8, sum(ANGLEtotal)
#row9 = cellA9, cellB9, cellC9, cellD9, sum(MAINStotal)
#row10 = cellA10, cellB10, cellC10, cellD10, sum(TEEStotal)
#row11 = cellA11, cellB11, cellC11, cellD11, sum(PANHEADStotal)
#row12 = cellA12, cellB12, cellC12, cellD12, sum(PINtotal)
#row13 = cellA13, cellB13, cellC13, cellD13, sum(SCREWStotal)

ListA = [
cellA1,
cellA2,
cellA3,
cellA4,
cellA5,
cellA6,
cellA7,
cellA8,
cellA9,
cellA10,
cellA11,
cellA12,
cellB1,
cellB2,
cellB3,
cellB4,
cellB5,
cellB6,
cellB7,
cellB8,
cellB9,
cellB10,
cellB11,
cellB12,
cellC1,
cellC2,
cellC3,
cellC4,
cellC5,
cellC6,
cellC7,
cellC8,
cellC9,
cellC10,
cellC11,
cellC12,
cellD1,
cellD2,
cellD3,
cellD4,
cellD5,
cellD6,
cellD7,
cellD8,
cellD9,
cellD10,
cellD11,
cellD12,
cellE1,
cellE2,
cellE3,
cellE4,
cellE5,
cellE6,
cellE7,
cellE8,
cellE9,
cellE10,
cellE11,
cellE12,
cellF1,
cellF2,
cellF3,
cellF4,
cellF5,
cellF6,
cellF7,
cellF8,
cellF9,
cellF10,
cellF11,
cellF12,
]

for xy in ListA:
    if xy == '5/8 FIRECODE':
        cellnumber =
        cellstring = str(xy)
        cellstring =
        FIRECODEqty =
     print(xy)
#print(row1)
#print(row2)
#print(row3)
#print(row4)
#print(row5)
#print(row6)
#print(row7)
#print(row8)
#print(row9)
#print(row10)
#print(row11)
#print(row12)
#print(row13)

#    cellA2 = sh.cell(row=2, column=1).value
#    cellC2 = sh.cell(row=2, column=3).value
#    cellA3 = sh.cell(row=3, column=1).value
#    cellA4 = sh.cell(row=4, column=1).value
#    cellA5 = sh.cell(row=5, column=1).value
#    cellA6 = sh.cell(row=6, column=1).value
#    cellA7 = sh.cell(row=7, column=1).value
#    cellA8 = sh.cell(row=8, column=1).value
#    cellA9 = sh.cell(row=9, column=1).value
#    if cellA2 == "Hat Channel" and cellC2 == '1 1/2" 25 GA':
#        qty = sh.cell(row=2, column=5).value
#        totals.append(qty)
#    if path =

    #else:
    #    print("Not Correct")
#print(sum(totals))

